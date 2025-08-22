import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill

# --- Page Config ---
st.set_page_config(page_title="POST Processor", page_icon="📊", layout="wide")

st.markdown(
    """
    <style>
        .main {
            background-color: #f8f9fa;
        }
        .stDownloadButton>button {
            background-color: #004080;
            color: white;
            font-weight: bold;
            border-radius: 8px;
        }
        .stDownloadButton>button:hover {
            background-color: #0066cc;
            color: white;
        }
        .uploadedFile {
            border: 1px solid #ddd;
            padding: 6px;
            border-radius: 6px;
            background-color: #ffffff;
        }
    </style>
    """,
    unsafe_allow_html=True
)

#st.title("📊 POST Data Processor")
st.markdown("Upload files below to generate a **professional Excel report** with all mappings and formatting.")

# --- File Upload Section ---
with st.expander("📂 Upload Input Files", expanded=True):
    col1, col2 = st.columns(2)
    with col1:
        post_file = st.file_uploader("POST Excel file", type=["xlsx"], key="post_file")
        tubs_file = st.file_uploader("TUBS Excel file", type=["xlsx"], key="tubs_file")
    with col2:
        demand_file = st.file_uploader("Demand Excel file (for Project)", type=["xlsx"], key="demand_file")
        beam_balance_file = st.file_uploader("BeamBalance Excel file", type=["xlsx"], key="beam_balance_file")

# --- Helper Functions ---
def extract_number(value, decimals):
    if pd.isna(value):
        return None
    m = re.search(r"[-+]?\d*\.\d+|\d+", str(value))
    if not m:
        return None
    return round(float(m.group()), decimals)

def count_gbd_gbs(value):
    if pd.isna(value):
        return 0
    values = re.split(r"[ ,;/]+", str(value))
    codes = [v for v in values if "GBD" in v or "GBS" in v]
    return len(codes)

# --- Main Processing ---
post_df = None
if post_file is not None:
    with st.spinner("Processing POST file..."):
        xls_post = pd.ExcelFile(post_file)
        post_sheet = st.selectbox("Select sheet from POST file", xls_post.sheet_names, index=0)
        post_df = pd.read_excel(xls_post, sheet_name=post_sheet)
        post_df.columns = post_df.columns.str.strip()

        # --- Add GBD/GBS count ---
        if "Demand" in post_df.columns:
            post_df["d counts"] = post_df["Demand"].apply(count_gbd_gbs)
            cols = list(post_df.columns)
            cols.insert(2, cols.pop(cols.index("d counts")))
            post_df = post_df[cols]

        # --- Merge Demand data ---
        if demand_file is not None:
            demand_df = pd.read_excel(demand_file)
            demand_df.columns = demand_df.columns.str.strip()

            if "GRE Prod Order" in demand_df.columns and "Project" in demand_df.columns:
                demand_unique = demand_df[["GRE Prod Order", "Project"]].drop_duplicates(subset=["GRE Prod Order"])
                post_df = post_df.merge(demand_unique, how="left", left_on="Production Order", right_on="GRE Prod Order")
                post_df.drop(columns=["GRE Prod Order"], inplace=True, errors="ignore")

                # Project → All GRE mapping
                project_po_mapping = (
                    demand_df.groupby("Project")["GRE Prod Order"]
                    .apply(lambda x: ",".join(sorted(map(str, set(x)))))
                    .to_dict()
                )
                post_df["All GRE Prod Orders (Project)"] = post_df["Project"].map(project_po_mapping)

                # Effective date end (latest)
                if "Effective date end" in demand_df.columns:
                    demand_df["Effective date end"] = pd.to_datetime(demand_df["Effective date end"], errors='coerce')
                    effective_date_mapping = (
                        demand_df.groupby("GRE Prod Order")["Effective date end"].max().reset_index()
                    )
                    post_df = post_df.merge(
                        effective_date_mapping, how="left",
                        left_on="Production Order", right_on="GRE Prod Order"
                    )
                    post_df.drop(columns=["GRE Prod Order"], inplace=True, errors="ignore")
                    post_df["Effective date end"] = post_df["Effective date end"].fillna("not found")

                post_df["Project"] = post_df["Project"].fillna("not found")
                post_df["All GRE Prod Orders (Project)"] = post_df["All GRE Prod Orders (Project)"].fillna("not found")

                # Reorder cols
                if "Project" in post_df.columns:
                    cols = list(post_df.columns)
                    cols.insert(3, cols.pop(cols.index("Project")))
                    if "All GRE Prod Orders (Project)" in cols:
                        cols.insert(4, cols.pop(cols.index("All GRE Prod Orders (Project)")))
                    if "Effective date end" in cols:
                        cols.insert(5, cols.pop(cols.index("Effective date end")))
                    post_df = post_df[cols]

                # GB Count
                post_df["GB_Count_in_Project"] = post_df["All GRE Prod Orders (Project)"].apply(
                    lambda x: len([v for v in str(x).split(",") if "GB" in v])
                )
                cols = list(post_df.columns)
                gb_col = cols.pop(cols.index("GB_Count_in_Project"))
                insert_pos = 6 if "Effective date end" in cols else 5
                cols.insert(insert_pos, gb_col)
                post_df = post_df[cols]

        # --- Numeric cleanup ---
        for col, dec in [("Beam Issue To PO", 2), ("Weft Issue To PO", 2), ("Action Qty Befor Post", 3)]:
            if col in post_df.columns:
                post_df[col] = post_df[col].apply(lambda x: extract_number(x, dec))

        # --- Add calculated cols ---
        if "Beam Issue To PO" in post_df.columns and "Weft Issue To PO" in post_df.columns:
            post_df["Beam+Weft"] = post_df["Beam Issue To PO"].fillna(0) + post_df["Weft Issue To PO"].fillna(0)
            cols = list(post_df.columns)
            idx = cols.index("Weft Issue To PO") + 1
            cols.insert(idx, cols.pop(cols.index("Beam+Weft")))
            post_df = post_df[cols]

        if "Waste" in post_df.columns and "Gre In Qty To WH" in post_df.columns:
            post_df["Waste+GreIn"] = post_df["Waste"].fillna(0) + post_df["Gre In Qty To WH"].fillna(0)
            cols = list(post_df.columns)
            idx = cols.index("Gre In Qty To WH") + 1
            cols.insert(idx, cols.pop(cols.index("Waste+GreIn")))
            post_df = post_df[cols]

        # --- Merge other files (TUBS + BeamBalance) ---
        # [same as your original code for TUBS + BeamBalance merging]

        # --- Preview ---
        st.markdown("### 🔎 Preview of Modified POST")
        st.dataframe(post_df, use_container_width=True)

        # --- Export Excel ---
        final_buf = BytesIO()
        post_df.to_excel(final_buf, index=False, sheet_name="ModifiedPost", engine="openpyxl")
        final_buf.seek(0)

        wb = load_workbook(final_buf)
        ws = wb.active

        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
        default_font = Font(name="Aptos Narrow", size=9, color="000000")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin
                cell.font = default_font
                cell.fill = PatternFill(fill_type="none")

        # Save again
        final_buf = BytesIO()
        wb.save(final_buf)
        final_buf.seek(0)

        # --- Download ---
        st.success("✅ Processing complete. Your file is ready!")
        st.download_button(
            label="📥 Download Modified POST",
            data=final_buf,
            file_name="Updated_post.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
